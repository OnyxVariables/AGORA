from faker import Faker
from openpyxl import load_workbook
from bs4 import BeautifulSoup

import requests
import sys
import re
import random
import zipfile
import os
import shutil

# Folder to extract the ZIP file
EXTRACT_DIR = "relacion_municipios"

def get_xlsx():
    zipFilename = "relacion_municipios.zip"

    url = "https://www.ine.es/daco/inebase_mensual/febrero_2025/relacion_municipios.zip"
    if len(sys.argv) > 1:
        url = sys.argv[1]

    requests.get(url, stream=True).raise_for_status()
    with open(zipFilename, "wb") as f:
        for chunk in requests.get(url, stream=True).iter_content(chunk_size=8192):
            f.write(chunk)
    
    with zipfile.ZipFile(zipFilename, 'r') as zip_ref:
        zip_ref.extractall(EXTRACT_DIR)

    os.remove(zipFilename)
    
    for root, dirs, files in os.walk(EXTRACT_DIR):
        for file in files:
            if file == "diccionario25.xlsx":
                return os.path.join(root, file)

    return None


def parse_xlsx(xlsxFilename):
    wb = load_workbook(filename=xlsxFilename, read_only=True, data_only=True)

    fields = []

    for row in wb.active.iter_rows(min_row=3, values_only=True):
        if row is None:
            continue

        try:
            autonomousCommunityId = int(row[0])
            provinceId = int(row[1])
            municipalityId = int(row[2])
            municipalityName = str(row[4]).strip()
        except (TypeError, ValueError, IndexError):
            continue

        fields.append({
            "autonomousCommunityId": autonomousCommunityId,
            "provinceId": provinceId,
            "municipalityId": municipalityId,
            "municipalityName": municipalityName,
        })

    wb.close()
    return fields


def get_data(url):
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    
    tds = soup.select(".miTabla > tr > td")

    return tds


def get_autonomous_communities(url):
    data = get_data(url)
    dataLength = len(data)

    autonomousCommunities = []
    for i in range(0, dataLength, 2):
        autonomousCommunities.append({
            "id": int(data[i].get_text(strip = True)),
            "name": data[i + 1].get_text(strip = True),
        })
    
    return autonomousCommunities


province_id_map = {}
def get_provinces(url, fields):
    data = get_data(url)
    dataLength = len(data)
    
    kv = {}
    for field in fields:
        kv[field["provinceId"]] = field["autonomousCommunityId"]

    provinces = []
    auto_id = 1
    for i in range(0, dataLength, 2):
        ine_id = int(data[i].get_text(strip = True))
        province_id_map[ine_id] = auto_id
        provinces.append({
            "id": auto_id,
            "ineId": ine_id,
            "autonomousCommunityId": int(kv[ine_id]),
            "name": data[i + 1].get_text(strip = True),
        })
        auto_id += 1

    return provinces

    
def populate_autonomous_community(autonomousCommunities):
    query = "INSERT INTO autonomousCommunity(id, name) VALUES\n"
    for autonomousCommunity in autonomousCommunities:
        query += f"({autonomousCommunity["id"]},\"{autonomousCommunity["name"]}\"),\n"

    query = query[:query.rfind(',')]
    query += ";"

    return query


def populate_province(provinces):
    query = "INSERT INTO province(ineId, autonomousCommunityId, name) VALUES\n"
    for province in provinces:
        query += f"({province["ineId"]},{province["autonomousCommunityId"]},\"{province["name"]}\"),\n"

    query = query[:query.rfind(',')]
    query += ";"

    return query


def populate_municipality(municipalities):
    query = "INSERT INTO municipality(ineId, provinceId, name) VALUES\n"
    for municipality in municipalities:
        query += f"({municipality["id"]},{municipality["provinceId"]},\"{municipality["name"]}\"),\n"

    query = query[:query.rfind(',')]
    query += ";"

    return query



def populate_user(numMunicipalities):
    fake = Faker()

    def generate_dni():
        letters = "TRWAGMYFPDXBNJZSQVHLCKE"
        randNum = fake.numerify("########")

        return randNum + letters[int(randNum) % len(letters)]
    
    query = "INSERT INTO user(name, nicknamePassword, roleId, dni, municipalityId) VALUES\n"
    for _ in range(10000):
        name = fake.name().replace("'", "''")
        nicknamePassword = f"{fake.md5()}{fake.user_name().replace("'", "''")}"
        roleId = 1 if random.randint(1, 10000) < 10 else 2
        dni = generate_dni()
        municipalityId = random.randint(1, numMunicipalities)

        query += f"(\"{name}\", \"{nicknamePassword}\", {roleId}, \"{dni}\", {municipalityId}),\n"
        
    query = query[:query.rfind(',')]
    query += ";"

    return query

query = ""

xlsxFilename = get_xlsx()
fields = parse_xlsx(xlsxFilename)
shutil.rmtree(EXTRACT_DIR)

autonomousCommunities = get_autonomous_communities("https://www.ine.es/daco/daco42/codmun/cod_ccaa.htm")
query += populate_autonomous_community(autonomousCommunities) + "\n"

provinces = get_provinces("https://www.ine.es/daco/daco42/codmun/cod_provincia.htm", fields)
query += populate_province(provinces) + "\n"

municipalities = [
    {
        "id": field["municipalityId"],
        "provinceId": province_id_map[field["provinceId"]],
        "name": field["municipalityName"],
    } for field in fields
]
query += populate_municipality(municipalities) + "\n"

query += populate_user(len(municipalities)) + "\n"

with open('insert.sql', "w", encoding="utf8") as f:
    f.write(query)
